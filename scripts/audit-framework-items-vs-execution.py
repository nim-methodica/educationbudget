import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
STATE_FILE = ROOT / ".tmp" / "state-current.json"
MAY_REPORT = ROOT / ".tmp" / "may-framework-update.xlsx"
OUT_FILE = ROOT / "outputs" / "framework-items-audit.json"

REG_COLUMNS = {
    "92": ("AY", "AZ", "BI", "BJ"),
    "46": ("BA", "BB", "BK", "BL"),
    "27": ("BC", "BD", "BM", "BN"),
    "73": ("BE", "BF", "BO", "BP"),
}


def n(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("₪", "").strip()
    if not text or text in {"-", "–"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def s(value):
    return str(value).strip() if value is not None else ""


def load_state():
    return json.loads(STATE_FILE.read_text(encoding="utf-8-sig"))


def extract_report_items():
    wb = load_workbook(MAY_REPORT, data_only=True)
    ws = wb[wb.sheetnames[1]]
    by_reg = {reg: {} for reg in REG_COLUMNS}
    for row_idx in range(1, ws.max_row + 1):
        code = s(ws[f"A{row_idx}"].value)
        name = s(ws[f"B{row_idx}"].value)
        unit_cost = n(ws[f"C{row_idx}"].value)
        if not code or not any(ch.isdigit() for ch in code) or code.startswith("סה"):
            continue
        for reg, (frame_qty_col, frame_amount_col, exec_qty_col, exec_amount_col) in REG_COLUMNS.items():
            frame_qty = n(ws[f"{frame_qty_col}{row_idx}"].value)
            frame_amount = n(ws[f"{frame_amount_col}{row_idx}"].value)
            exec_qty = n(ws[f"{exec_qty_col}{row_idx}"].value)
            exec_amount = n(ws[f"{exec_amount_col}{row_idx}"].value)
            if any([frame_qty, frame_amount, exec_qty, exec_amount]):
                by_reg[reg][code] = {
                    "code": code,
                    "name": name,
                    "unitCost": unit_cost,
                    "reportFrameworkQty": frame_qty,
                    "reportFrameworkAmount": frame_amount,
                    "reportExecutionQty": exec_qty,
                    "reportExecutionAmount": exec_amount,
                }
    return by_reg


def order_quantity_maps(regulation):
    reserved = defaultdict(float)
    paid = defaultdict(float)
    orders = defaultdict(list)
    for order in regulation.get("projectOrders", []):
        is_closed = order.get("status") == "closed"
        for line in order.get("lines", []):
            code = s(line.get("code"))
            qty = n(line.get("quantity"))
            if not code:
                continue
            reserved[code] += qty
            paid_qty = qty if is_closed else n(line.get("collectedQuantity"))
            paid[code] += paid_qty
            orders[code].append({
                "orderNumber": order.get("orderNumber"),
                "projectName": order.get("projectName") or order.get("title"),
                "status": order.get("status"),
                "quantity": qty,
                "paidQuantity": paid_qty,
            })
    return reserved, paid, orders


def close(a, b, tolerance=0.01):
    return abs((a or 0) - (b or 0)) <= tolerance


def item_sort_key(code):
    parts = []
    for part in str(code).split("."):
        if part.isdigit():
            parts.append((0, int(part)))
        else:
            parts.append((1, part))
    return parts


def build_audit():
    state = load_state()
    framework = state["frameworks"][0]
    report = extract_report_items()
    audit = []
    issues = []

    for regulation in framework.get("regulations", []):
        reg = str(regulation.get("number"))
        if reg not in REG_COLUMNS:
            continue
        reserved, paid, orders = order_quantity_maps(regulation)
        system_items = {s(item.get("code")): item for item in regulation.get("items", [])}
        codes = sorted(
            set(system_items) | set(report[reg]) | set(reserved) | set(paid),
            key=item_sort_key
        )
        for code in codes:
            system_item = system_items.get(code, {})
            report_item = report[reg].get(code, {})
            system_qty = n(system_item.get("approvedQuantity"))
            report_qty = n(report_item.get("reportFrameworkQty"))
            reserved_qty = reserved[code]
            paid_qty = paid[code]
            execution_qty = n(report_item.get("reportExecutionQty"))
            unit_cost = n(system_item.get("unitCost") or report_item.get("unitCost"))
            row_issues = []
            if not close(system_qty, report_qty):
                row_issues.append("פער מסגרת מול דוח")
            if reserved_qty > system_qty + 0.01:
                row_issues.append("ניצול בהזמנות מעל מסגרת")
            if paid_qty > reserved_qty + 0.01:
                row_issues.append("שולם מעל ניצול")
            if execution_qty > reserved_qty + 0.01:
                row_issues.append("ביצוע מצטבר מעל הזמנות")
            if row_issues:
                issues.append({
                    "regulation": reg,
                    "code": code,
                    "name": system_item.get("name") or report_item.get("name"),
                    "systemFrameworkQty": system_qty,
                    "reportFrameworkQty": report_qty,
                    "reservedQty": reserved_qty,
                    "paidQty": paid_qty,
                    "executionQty": execution_qty,
                    "issues": row_issues,
                    "orders": orders[code],
                })
            audit.append({
                "regulation": reg,
                "code": code,
                "name": system_item.get("name") or report_item.get("name"),
                "unitCost": unit_cost,
                "systemFrameworkQty": system_qty,
                "reportFrameworkQty": report_qty,
                "reservedQty": reserved_qty,
                "paidQty": paid_qty,
                "executionQty": execution_qty,
                "issues": row_issues,
            })
    return {"audit": audit, "issues": issues}


if __name__ == "__main__":
    result = build_audit()
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "rows": len(result["audit"]),
        "issues": len(result["issues"]),
        "byRegulation": {
            reg: sum(1 for row in result["issues"] if row["regulation"] == reg)
            for reg in REG_COLUMNS
        }
    }, ensure_ascii=True))
