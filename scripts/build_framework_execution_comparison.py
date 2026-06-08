from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / ".tmp" / "may-framework-update.xlsx"
OUTPUT = ROOT / "outputs" / "השוואת-מסגרת-מול-ביצוע-מצטבר-מאי-2026.xlsx"


REGULATIONS = {
    "92": {
        "name": "מינהל חדשנות וטכנולוגיה",
        "framework_qty": "AY",
        "framework_amount": "AZ",
        "execution_qty": "BI",
        "execution_amount": "BJ",
    },
    "46": {
        "name": "מזכירות פדגוגית / מזה\"פ",
        "framework_qty": "BA",
        "framework_amount": "BB",
        "execution_qty": "BK",
        "execution_amount": "BL",
    },
    "27": {
        "name": "ישראל ריאלית / STEM",
        "framework_qty": "BC",
        "framework_amount": "BD",
        "execution_qty": "BM",
        "execution_amount": "BN",
    },
    "73": {
        "name": "בינה מלאכותית",
        "framework_qty": "BE",
        "framework_amount": "BF",
        "execution_qty": "BO",
        "execution_amount": "BP",
    },
}


DETAIL_HEADERS = [
    "סעיף",
    "שם פריט",
    "עלות ללא מע\"מ",
    "כמות במסגרת",
    "סכום במסגרת ללא מע\"מ",
    "כמות ביצוע מצטבר",
    "סכום ביצוע מצטבר ללא מע\"מ",
    "יתרת כמות",
    "יתרת סכום ללא מע\"מ",
    "סטטוס",
]


SUMMARY_HEADERS = [
    "תקנה",
    "שם תקנה",
    "עמודות מקור",
    "כמות במסגרת",
    "סכום במסגרת ללא מע\"מ",
    "כמות ביצוע מצטבר",
    "סכום ביצוע מצטבר ללא מע\"מ",
    "יתרת כמות",
    "יתרת סכום ללא מע\"מ",
    "מספר חריגות",
]


def to_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("₪", "").strip()
    if not text or text in {"-", "–"}:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def clean_text(value):
    return str(value).strip() if value is not None else ""


def status_for(framework_qty, execution_qty, framework_amount, execution_amount):
    if framework_qty == 0 and execution_qty > 0:
        return "ביצוע ללא מסגרת"
    if execution_qty > framework_qty:
        return "חריגה"
    if framework_qty > 0 and execution_qty == 0:
        return "לא נוצל"
    if framework_qty > 0 and execution_qty == framework_qty:
        return "נוצל מלא"
    if framework_qty > execution_qty:
        return "יש יתרה"
    if framework_amount == 0 and execution_amount > 0:
        return "ביצוע ללא מסגרת"
    return "לבדיקה"


def fmt_number(value):
    if value is None:
        return 0
    rounded = round(value, 2)
    if rounded == int(rounded):
        return int(rounded)
    return rounded


def style_sheet(ws, freeze="A2"):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    header_fill = PatternFill("solid", fgColor="EAF3EF")
    border = Border(bottom=Side(style="thin", color="D9E3DF"))

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Assistant", size=11)
            cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
            cell.border = border

    for cell in ws[1]:
        cell.font = Font(name="Assistant", bold=True, size=11, color="173E39")
        cell.fill = header_fill

    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def apply_number_formats(ws, money_cols, qty_cols):
    for row in ws.iter_rows(min_row=2):
        for col_idx in money_cols:
            row[col_idx - 1].number_format = '#,##0.00 "₪";-#,##0.00 "₪"'
        for col_idx in qty_cols:
            row[col_idx - 1].number_format = '#,##0.##'


def extract_rows(source_ws, reg_code, cfg):
    rows = []
    for row_idx in range(1, source_ws.max_row + 1):
        item_code = clean_text(source_ws[f"A{row_idx}"].value)
        item_name = clean_text(source_ws[f"B{row_idx}"].value)
        unit_cost = to_number(source_ws[f"C{row_idx}"].value)

        if not item_code or item_code.startswith("סה"):
            continue
        if not any(ch.isdigit() for ch in item_code):
            continue

        framework_qty = to_number(source_ws[f"{cfg['framework_qty']}{row_idx}"].value)
        framework_amount = to_number(source_ws[f"{cfg['framework_amount']}{row_idx}"].value)
        execution_qty = to_number(source_ws[f"{cfg['execution_qty']}{row_idx}"].value)
        execution_amount = to_number(source_ws[f"{cfg['execution_amount']}{row_idx}"].value)

        if not any([framework_qty, framework_amount, execution_qty, execution_amount]):
            continue

        balance_qty = framework_qty - execution_qty
        balance_amount = framework_amount - execution_amount

        rows.append(
            {
                "regulation": reg_code,
                "item_code": item_code,
                "item_name": item_name,
                "unit_cost": unit_cost,
                "framework_qty": framework_qty,
                "framework_amount": framework_amount,
                "execution_qty": execution_qty,
                "execution_amount": execution_amount,
                "balance_qty": balance_qty,
                "balance_amount": balance_amount,
                "status": status_for(framework_qty, execution_qty, framework_amount, execution_amount),
            }
        )
    return rows


def build_workbook():
    if not SOURCE.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE}")

    source_wb = load_workbook(SOURCE, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[1]]

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "סיכום"
    summary_ws.append(SUMMARY_HEADERS)

    for reg_code, cfg in REGULATIONS.items():
        rows = extract_rows(source_ws, reg_code, cfg)

        detail_ws = wb.create_sheet(reg_code)
        detail_ws.append(DETAIL_HEADERS)
        for item in rows:
            detail_ws.append(
                [
                    item["item_code"],
                    item["item_name"],
                    fmt_number(item["unit_cost"]),
                    fmt_number(item["framework_qty"]),
                    fmt_number(item["framework_amount"]),
                    fmt_number(item["execution_qty"]),
                    fmt_number(item["execution_amount"]),
                    fmt_number(item["balance_qty"]),
                    fmt_number(item["balance_amount"]),
                    item["status"],
                ]
            )

        style_sheet(detail_ws)
        set_widths(detail_ws, [12, 42, 16, 16, 22, 20, 25, 14, 22, 18])
        apply_number_formats(detail_ws, money_cols=[3, 5, 7, 9], qty_cols=[4, 6, 8])

        totals = {
            "framework_qty": sum(item["framework_qty"] for item in rows),
            "framework_amount": sum(item["framework_amount"] for item in rows),
            "execution_qty": sum(item["execution_qty"] for item in rows),
            "execution_amount": sum(item["execution_amount"] for item in rows),
            "balance_qty": sum(item["balance_qty"] for item in rows),
            "balance_amount": sum(item["balance_amount"] for item in rows),
            "exceptions": sum(1 for item in rows if item["status"] in {"חריגה", "ביצוע ללא מסגרת"}),
        }

        source_cols = (
            f"{cfg['framework_qty']}+{cfg['execution_qty']} "
            f"(סכומים: {cfg['framework_amount']}+{cfg['execution_amount']})"
        )
        summary_ws.append(
            [
                reg_code,
                cfg["name"],
                source_cols,
                fmt_number(totals["framework_qty"]),
                fmt_number(totals["framework_amount"]),
                fmt_number(totals["execution_qty"]),
                fmt_number(totals["execution_amount"]),
                fmt_number(totals["balance_qty"]),
                fmt_number(totals["balance_amount"]),
                totals["exceptions"],
            ]
        )

    style_sheet(summary_ws)
    set_widths(summary_ws, [10, 30, 32, 18, 24, 22, 28, 18, 24, 14])
    apply_number_formats(summary_ws, money_cols=[5, 7, 9], qty_cols=[4, 6, 8, 10])

    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    output = build_workbook()
    print(str(output).encode("unicode_escape").decode("ascii"))
