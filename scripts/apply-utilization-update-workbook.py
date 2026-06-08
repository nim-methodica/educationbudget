import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "app-data.json"
OUTPUT_DIR = ROOT / "outputs"
BACKUP_DIR = ROOT / ".tmp" / "backups"
WORKBOOK_NAME = "\u05e2\u05d3\u05db\u05d5\u05df-\u05e0\u05d9\u05e6\u05d5\u05dc-\u05e9\u05d5\u05e8\u05d5\u05ea-\u05d4\u05d6\u05de\u05e0\u05d5\u05ea-\u05de\u05de\u05d5\u05d9\u05df.xlsx"
REPORT_FILE = OUTPUT_DIR / "utilization-update-apply-report.json"


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_number(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    return float(text) if text else 0


def normalized_code(value):
    text = as_text(value)
    return text[:-2] if text.endswith(".0") else text


def is_marked(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
    color = fill.fgColor
    rgb = str(color.rgb or "")
    return rgb.upper().endswith("FFFF00")


def load_state():
    return json.loads(DATA_FILE.read_text(encoding="utf-8-sig"))


def save_state(state):
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(DATA_FILE, BACKUP_DIR / f"app-data-before-utilization-update-{stamp}.json")
    DATA_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def find_workbook():
    path = OUTPUT_DIR / WORKBOOK_NAME
    if path.exists():
        return path
    matches = list(OUTPUT_DIR.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError("No xlsx files were found in outputs")
    return max(matches, key=lambda entry: entry.stat().st_mtime)


def read_workbook_rows(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = []
    for excel_row in range(2, sheet.max_row + 1):
        values = [sheet.cell(excel_row, col).value for col in range(1, 10)]
        if not any(value is not None and str(value).strip() for value in values):
            continue
        marked_cols = [col for col in range(1, min(sheet.max_column, 10) + 1) if is_marked(sheet.cell(excel_row, col))]
        rows.append({
            "excelRow": excel_row,
            "serial": int(as_number(values[0])) if values[0] not in (None, "") else None,
            "regulation": as_text(values[1]),
            "orderNumber": as_text(values[2]),
            "projectName": as_text(values[3]),
            "status": as_text(values[4]),
            "code": normalized_code(values[5]),
            "name": as_text(values[6]),
            "quantity": as_number(values[7]),
            "utilizedQuantity": as_number(values[8]),
            "markedCols": marked_cols,
            "marked": bool(marked_cols),
        })
    return rows


def iter_orders(state):
    for framework in state.get("frameworks", []):
        for regulation in framework.get("regulations", []):
            for order in regulation.get("projectOrders", []):
                yield framework, regulation, order


def framework_item_lookup(state):
    lookup = {}
    for framework in state.get("frameworks", []):
        for regulation in framework.get("regulations", []):
            reg_number = as_text(regulation.get("number"))
            for item in regulation.get("items", []):
                lookup[(reg_number, normalized_code(item.get("code")))] = item
    return lookup


def existing_name_for(order, code, occurrence_index):
    matches = [line for line in order.get("lines", []) if normalized_code(line.get("code")) == code]
    if occurrence_index < len(matches):
        return as_text(matches[occurrence_index].get("name"))
    return ""


def existing_unit_cost_for(order, code, occurrence_index, item_lookup, regulation):
    matches = [line for line in order.get("lines", []) if normalized_code(line.get("code")) == code]
    if occurrence_index < len(matches):
        return as_number(matches[occurrence_index].get("unitCost"))
    item = item_lookup.get((as_text(regulation.get("number")), code))
    return as_number(item.get("unitCost")) if item else 0


def line_from_workbook_row(row, order, regulation, item_lookup, occurrence_index):
    code = row["code"]
    item = item_lookup.get((as_text(regulation.get("number")), code))
    name = row["name"] or existing_name_for(order, code, occurrence_index) or as_text((item or {}).get("name"))
    return {
        "code": code,
        "name": name,
        "quantity": row["quantity"],
        "unitCost": existing_unit_cost_for(order, code, occurrence_index, item_lookup, regulation),
        "utilizedQuantity": row["utilizedQuantity"],
    }


def current_export_refs(state):
    rows = []
    for _, regulation, order in iter_orders(state):
        for line_index, line in enumerate(order.get("lines", [])):
            if order.get("orderNumber") == "25-019" and normalized_code(line.get("code")) == "\u05e1\u05d9\u05db\u05d5\u05dd":
                continue
            rows.append({
                "orderNumber": as_text(order.get("orderNumber")),
                "regulation": as_text(regulation.get("number")),
                "code": normalized_code(line.get("code")),
                "lineIndex": line_index,
                "order": order,
                "regulationObject": regulation,
            })
    rows.sort(key=lambda row: (row["orderNumber"], row["regulation"], row["code"]))
    return {index: row for index, row in enumerate(rows, start=1)}


def apply_updates(state, workbook_rows):
    order_index = {as_text(order.get("orderNumber")): (framework, regulation, order) for framework, regulation, order in iter_orders(state)}
    item_lookup = framework_item_lookup(state)
    refs = current_export_refs(state)
    grouped_rows = defaultdict(list)
    for row in workbook_rows:
        if row["orderNumber"]:
            grouped_rows[row["orderNumber"]].append(row)

    report = {
        "workbookRows": len(workbook_rows),
        "markedRows": sum(1 for row in workbook_rows if row["marked"]),
        "ordersTouched": [],
        "changes": [],
        "warnings": [],
    }

    for order_number, rows_for_order in grouped_rows.items():
        marked_rows = [row for row in rows_for_order if row["marked"]]
        if not marked_rows:
            continue
        if order_number not in order_index:
            report["warnings"].append({"orderNumber": order_number, "issue": "order not found"})
            continue

        _, regulation, order = order_index[order_number]
        report["ordersTouched"].append(order_number)
        structure_marked = any(any(col in row["markedCols"] for col in (3, 4, 6, 7, 8)) for row in marked_rows)
        replace_order_lines = structure_marked and all(row["marked"] for row in rows_for_order)

        if replace_order_lines:
            occurrences = Counter()
            old_lines = order.get("lines", [])
            new_lines = []
            for row in rows_for_order:
                code = row["code"]
                occurrence_index = occurrences[code]
                occurrences[code] += 1
                new_lines.append(line_from_workbook_row(row, order, regulation, item_lookup, occurrence_index))
            order["lines"] = new_lines
            order["totalWithoutVat"] = sum(as_number(line.get("quantity")) * as_number(line.get("unitCost")) for line in new_lines)
            report["changes"].append({
                "orderNumber": order_number,
                "type": "replace-lines",
                "oldLineCount": len(old_lines),
                "newLineCount": len(new_lines),
            })
            continue

        workbook_occurrences = Counter()
        for row in rows_for_order:
            code = row["code"]
            occurrence_index = workbook_occurrences[code]
            workbook_occurrences[code] += 1
            if not row["marked"]:
                continue

            line = None
            ref = refs.get(row["serial"])
            if ref and ref["orderNumber"] == order_number:
                line = order.get("lines", [])[ref["lineIndex"]]
            if line is None:
                matches = [entry for entry in order.get("lines", []) if normalized_code(entry.get("code")) == code]
                if occurrence_index < len(matches):
                    line = matches[occurrence_index]
            if line is None:
                line = line_from_workbook_row(row, order, regulation, item_lookup, occurrence_index)
                order.setdefault("lines", []).append(line)
                report["changes"].append({"orderNumber": order_number, "type": "add-line", "code": code})

            before = {
                "code": normalized_code(line.get("code")),
                "name": as_text(line.get("name")),
                "quantity": as_number(line.get("quantity")),
                "utilizedQuantity": line.get("utilizedQuantity"),
            }
            if 6 in row["markedCols"]:
                line["code"] = code
            if row["name"] and (7 in row["markedCols"] or not as_text(line.get("name"))):
                line["name"] = row["name"]
            if 8 in row["markedCols"]:
                line["quantity"] = row["quantity"]
            if 9 in row["markedCols"]:
                line["utilizedQuantity"] = row["utilizedQuantity"]
            after = {
                "code": normalized_code(line.get("code")),
                "name": as_text(line.get("name")),
                "quantity": as_number(line.get("quantity")),
                "utilizedQuantity": line.get("utilizedQuantity"),
            }
            if before != after:
                report["changes"].append({
                    "orderNumber": order_number,
                    "code": code,
                    "before": before,
                    "after": after,
                })

        order["totalWithoutVat"] = sum(as_number(line.get("quantity")) * as_number(line.get("unitCost")) for line in order.get("lines", []))

    report["missingNamesFilled"] = fill_missing_order_line_names(state)
    report["ordersTouched"] = sorted(set(report["ordersTouched"]))
    return report


def fill_missing_order_line_names(state):
    item_lookup = framework_item_lookup(state)
    count = 0
    for _, regulation, order in iter_orders(state):
        reg_number = as_text(regulation.get("number"))
        for line in order.get("lines", []):
            if as_text(line.get("name")):
                continue
            item = item_lookup.get((reg_number, normalized_code(line.get("code"))))
            if item and as_text(item.get("name")):
                line["name"] = as_text(item.get("name"))
                count += 1
    return count


def main():
    workbook_path = find_workbook()
    state = load_state()
    rows = read_workbook_rows(workbook_path)
    report = apply_updates(state, rows)
    save_state(state)
    report["workbook"] = str(workbook_path)
    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "markedRows": report["markedRows"],
        "ordersTouched": len(report["ordersTouched"]),
        "changes": len(report["changes"]),
        "warnings": len(report["warnings"]),
        "report": str(REPORT_FILE),
    }, ensure_ascii=True))


if __name__ == "__main__":
    main()
