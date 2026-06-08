import json
from pathlib import Path
from urllib.error import URLError
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
JSON_FILE = ROOT / "data" / "app-data.json"
OUTPUT = ROOT / "outputs" / "עדכון-ניצול-שורות-הזמנות-ממוין.xlsx"
API_STATE = "http://localhost:4235/api/state"


HEADERS = [
    "מס׳",
    "תקנה",
    "מספר הזמנה",
    "שם הזמנה",
    "סטטוס הזמנה",
    "סעיף",
    "שם פריט",
    "כמות פריטים בהזמנה",
    "נוצל",
]


def load_state():
    try:
        with urlopen(API_STATE, timeout=5) as response:
            return json.loads(response.read().decode("utf-8"))
    except (URLError, TimeoutError, json.JSONDecodeError):
        return json.loads(JSON_FILE.read_text(encoding="utf-8-sig"))


def status_label(status):
    if status == "closed":
        return "סגורה"
    if status == "draft":
        return "טיוטה"
    return "פתוחה"


def build_rows(state):
    rows = []
    for framework in state.get("frameworks", []):
        if not framework.get("isDefault") and framework.get("id") != state.get("defaultFrameworkId"):
            continue
        for regulation in framework.get("regulations", []):
            regulation_number = str(regulation.get("number", ""))
            for order in regulation.get("projectOrders", []):
                is_closed = order.get("status") == "closed"
                for line in order.get("lines", []):
                    if order.get("orderNumber") == "25-019" and str(line.get("code", "")).strip() == "סיכום":
                        continue
                    quantity = float(line.get("quantity") or 0)
                    rows.append([
                        regulation_number,
                        order.get("orderNumber", ""),
                        order.get("projectName") or order.get("title") or "",
                        status_label(order.get("status")),
                        str(line.get("code", "")),
                        line.get("name", ""),
                        quantity,
                        quantity if is_closed else 0,
                    ])
    rows.sort(key=lambda row: (str(row[1]), str(row[0]), str(row[4])))
    return [[index, *row] for index, row in enumerate(rows, start=1)]


def style_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="EAF3EF")
    border = Border(bottom=Side(style="thin", color="D9E3DF"))

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Assistant", size=11)
            cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
            cell.border = border

    for cell in ws[1]:
        cell.font = Font(name="Assistant", bold=True, size=11, color="173E39")
        cell.fill = header_fill

    widths = [8, 10, 15, 42, 14, 12, 46, 18, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.number_format = "#,##0.###"


def main():
    state = load_state()
    rows = build_rows(state)

    wb = Workbook()
    ws = wb.active
    ws.title = "שורות הזמנות"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    style_sheet(ws)

    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "ערך"])
    summary.append(["מספר שורות", len(rows)])
    summary.append(["מספר הזמנות", len({(row[1], row[2]) for row in rows})])
    summary.append(["סך כמות בהזמנות", sum(float(row[7] or 0) for row in rows)])
    summary.append(["סך נוצל ראשוני", sum(float(row[8] or 0) for row in rows)])
    style_sheet(summary)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(str(OUTPUT).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
