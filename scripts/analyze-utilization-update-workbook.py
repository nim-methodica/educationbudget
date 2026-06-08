import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "עדכון-ניצול-שורות-הזמנות-ממוין.xlsx"
OUT_JSON = ROOT / "outputs" / "utilization-update-yellow-rows.json"
OUT_MD = ROOT / "outputs" / "שורות-צהובות-עדכון-ניצול.md"


def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return False
    colors = []
    for color in (fill.fgColor, fill.start_color):
        rgb = getattr(color, "rgb", None)
        indexed = getattr(color, "indexed", None)
        if rgb:
            colors.append(str(rgb).upper())
        if indexed is not None:
            colors.append(f"INDEXED:{indexed}")
    return any(color in {"FFFFFF00", "FFFF00", "FFFFFF99", "FFFF99", "INDEXED:13"} for color in colors)


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def main():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["שורות הזמנות"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row_idx, col) for col in range(1, ws.max_column + 1)]
        yellow_cols = [headers[index] for index, cell in enumerate(cells) if is_yellow(cell)]
        if not yellow_cols:
            continue
        row = {headers[index]: clean(cell.value) for index, cell in enumerate(cells)}
        row["_excelRow"] = row_idx
        row["_yellowColumns"] = yellow_cols
        rows.append(row)

    OUT_JSON.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# שורות צהובות - עדכון ניצול",
        "",
        "| שורה | תקנה | הזמנה | סעיף | כמות בהזמנה | נוצל | עמודות צהובות |",
        "|---:|---|---|---|---:|---:|---|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join([
                str(row.get("_excelRow", "")),
                str(row.get("תקנה", "")),
                str(row.get("מספר הזמנה", "")),
                str(row.get("סעיף", "")),
                str(row.get("כמות פריטים בהזמנה", "")),
                str(row.get("נוצל", "")),
                ", ".join(str(col) for col in row.get("_yellowColumns", [])),
            ])
            + " |"
        )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(json.dumps({"yellowRows": len(rows), "json": str(OUT_JSON), "md": str(OUT_MD)}, ensure_ascii=True))


if __name__ == "__main__":
    main()
